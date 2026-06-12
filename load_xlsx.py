"""
One-time script: reads the 4 Excel files and appends their content
to system_prompt.txt so Claude has the actual data at context load.

Run: python load_xlsx.py
"""

import openpyxl
from pathlib import Path

FILES = [
    ("תוכניות_אימון_2024-2025.xlsx", "תוכניות עונה 2024-2025"),
    ("תוכניות_אימון_2025-2026.xlsx", "תוכניות עונה 2025-2026"),
    ("תרגילי_גודו_למבחן.xlsx", "טכניקות לפי חגורה"),
    ("מערך_אימון_ותרגילי_גודו.xlsx", "חימומים, משחקים, הפלות, ריתוקים, כוח"),
]

MAX_ROWS = 200  # limit per sheet to avoid huge context


def xlsx_to_text(path: str, label: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = [f"\n===== {label} ====="]
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        parts.append(f"\n-- גיליון: {sheet} --")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS:
                parts.append(f"  [... חתוך אחרי {MAX_ROWS} שורות]")
                break
            row_text = " | ".join(str(c) if c is not None else "" for c in row)
            if row_text.strip("|").strip():
                parts.append(row_text)
    return "\n".join(parts)


def main():
    base_prompt = Path("system_prompt.txt").read_text(encoding="utf-8")
    additions = []
    for filename, label in FILES:
        if Path(filename).exists():
            print(f"Loading {filename}...")
            additions.append(xlsx_to_text(filename, label))
        else:
            print(f"  ⚠️  {filename} not found — skipping")

    if additions:
        full = base_prompt + "\n" + "\n".join(additions)
        Path("system_prompt_with_data.txt").write_text(full, encoding="utf-8")
        print("✅  Saved system_prompt_with_data.txt")
        print("   Update bot.py to load this file instead of system_prompt.txt")
    else:
        print("No Excel files found. Place them next to this script and re-run.")


if __name__ == "__main__":
    main()
