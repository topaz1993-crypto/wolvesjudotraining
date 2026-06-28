"""
invoice4u_reader.py — קריאה וסיווג קובץ XLS/XLSX מ-invoice4u.
"""

import io, re
from pathlib import Path
from typing import Optional
from datetime import date as date_cls
import openpyxl

MONTH_HE = {
    '01': 'ינואר', '02': 'פברואר', '03': 'מרץ',    '04': 'אפריל',
    '05': 'מאי',   '06': 'יוני',   '07': 'יולי',    '08': 'אוגוסט',
    '09': 'ספטמבר','10': 'אוקטובר','11': 'נובמבר',  '12': 'דצמבר',
}

BELT_AMOUNT      = 60
LARGE_THRESHOLD  = 5000   # above this = club transfer
CLUB_XFER_NAMES  = {'עמותת ספורט כפר סירקין'}

# Amounts that map cleanly to a subscription type
KNOWN_MONTHLY = {200, 220, 280, 300, 400, 420, 440, 600}


def _parse_date(s: str) -> Optional[date_cls]:
    m = re.match(r'(\d{1,2})[/.](\d{1,2})(?:[/.](\d{2,4}))?', str(s).strip())
    if not m:
        return None
    d, mo = int(m.group(1)), int(m.group(2))
    y = int(m.group(3)) if m.group(3) else date_cls.today().year
    if y < 100:
        y += 2000
    try:
        return date_cls(y, mo, d)
    except ValueError:
        return None


def _extract_children(name: str) -> tuple:
    """
    "רן כהן (יואב)"          → parent="רן כהן",       children=["יואב"]
    "טלי גירו (ינאי ולביא)"  → parent="טלי גירו",      children=["ינאי","לביא"]
    "נלי אשכנזי"              → parent="נלי אשכנזי",    children=[]
    """
    m = re.search(r'\(([^)]+)\)', name)
    if m:
        parent = name[:m.start()].strip()
        inner  = m.group(1)
        # Split on " ו" (Hebrew "and") or comma
        children = re.split(r'\s+ו(?=[א-ת])|,\s*', inner)
        children = [c.strip() for c in children if c.strip()]
        return parent, children
    return name.strip(), []


def _classify(name: str, amount: int) -> str:
    """Return: 'monthly' | 'belt' | 'club_transfer' | 'other'"""
    if name in CLUB_XFER_NAMES or amount >= LARGE_THRESHOLD:
        return 'club_transfer'
    if amount == BELT_AMOUNT:
        return 'belt'
    if amount in KNOWN_MONTHLY:
        return 'monthly'
    if 100 <= amount <= 700:
        return 'monthly_unusual'  # possible partial / discount — flag to user
    return 'other'


def read_xls(source) -> list[dict]:
    """
    Read invoice4u XLS/XLSX file (path, bytes, or file-like object).
    Returns list of payment record dicts.
    """
    if isinstance(source, (str, Path)):
        with open(source, 'rb') as f:
            data = f.read()
    elif isinstance(source, bytes):
        data = source
    else:
        data = source.read()

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    TARGET = 'חשבונית מס קבלה'
    if TARGET not in wb.sheetnames:
        raise ValueError(f"לא נמצא גיליון '{TARGET}' בקובץ")

    sh = wb[TARGET]
    records = []

    for row in sh.iter_rows(min_row=4, values_only=True):
        date_val  = row[0]
        if not date_val:
            continue
        date_str = str(date_val).strip()
        if date_str in ('', 'סה"כ ₪'):
            continue

        invoice_no    = str(row[1] or '').strip()
        customer_name = str(row[2] or '').strip()
        customer_id   = str(row[4] or '').strip()

        try:
            amount = int(float(str(row[7] or 0)))
        except (ValueError, TypeError):
            amount = 0

        if not customer_name or amount <= 0:
            continue

        parsed      = _parse_date(date_str)
        month_num   = f"{parsed.month:02d}" if parsed else '??'
        month_he    = MONTH_HE.get(month_num, month_num)
        year        = parsed.year if parsed else date_cls.today().year

        parent_name, children = _extract_children(customer_name)

        records.append({
            'date':          date_str,
            'parsed_date':   parsed,
            'month':         month_he,
            'month_num':     month_num,
            'year':          year,
            'invoice_no':    invoice_no,
            'customer_name': customer_name,
            'customer_id':   customer_id,
            'amount':        amount,
            'type':          _classify(customer_name, amount),
            'parent_name':   parent_name,
            'children':      children,
        })

    return records


def available_months(records: list[dict]) -> list[str]:
    """Return unique 'Month YYYY' strings sorted chronologically."""
    seen = {}
    for r in records:
        key = f"{r['month']} {r['year']}"
        seen[key] = (r['year'], int(r['month_num']) if r['month_num'].isdigit() else 0)
    return [k for k, _ in sorted(seen.items(), key=lambda x: x[1])]


def filter_month(records: list[dict], month_he: str,
                 year: Optional[int] = None) -> list[dict]:
    """Return records for one Hebrew month (optionally filtered by year)."""
    return [
        r for r in records
        if r['month'] == month_he and (year is None or r['year'] == year)
    ]


def summarise(records: list[dict]) -> dict:
    """Count records by type and sum amounts."""
    by_type: dict[str, list] = {
        'monthly': [], 'monthly_unusual': [],
        'belt': [], 'club_transfer': [], 'other': [],
    }
    for r in records:
        by_type.setdefault(r['type'], []).append(r)
    return {
        'total':           len(records),
        'monthly':         by_type['monthly'],
        'monthly_unusual': by_type['monthly_unusual'],
        'belt':            by_type['belt'],
        'club_transfer':   by_type['club_transfer'],
        'other':           by_type['other'],
        'total_amount':    sum(r['amount'] for r in records),
    }
